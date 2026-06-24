"""
BOQ Excel writer — 원본 .xls 공내역서를 읽어 단가·금액을 채운 .xlsx로 저장한다.

흐름:
  1) 원본 .xls 시트별로 raw cell-by-cell 읽기 (xlrd)
  2) excel_recognizer로 분류해 item 행을 식별
  3) boq_pricing로 단가 매핑 → unit_price + amount 채움
  4) openpyxl로 새 .xlsx 작성 (헤더/합계행 그대로 유지)
  5) 시트별 work_group 소계와 전체 합계를 자동 추가

Run:
    python -X utf8 scripts/estimation/boq_excel_writer.py <xls_in> <xlsx_out>
"""
from __future__ import annotations
import argparse
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from boq_pricing import is_total_row, lookup_unit_price
from excel_recognizer import _classify_row, _coerce_number, _open, _is_header_row


THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="FFE2EFDA")
SUMMARY_FILL = PatternFill("solid", fgColor="FFFFF2CC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="FFD9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFFCE4D6")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

NUM_FMT = "#,##0"

# 출력 컬럼 (12개): 공종/번호/품명/제조사/모델/규격/단위/수량/횟수/단가/금액/단가출처
OUT_COLS = [
    ("공종", 10), ("번호", 6), ("품명", 28), ("제조사", 12),
    ("모델", 12), ("규격", 22), ("단위", 8), ("수량", 10),
    ("횟수", 8), ("단가", 12), ("금액", 14), ("단가출처", 22),
]


def _read_raw_rows(xls_path: Path, sheet: str, engine: str) -> list[list[str]]:
    df = pd.read_excel(xls_path, sheet_name=sheet, header=None, engine=engine)
    out: list[list[str]] = []
    for i in range(len(df)):
        row = df.iloc[i].fillna("").astype(str).tolist()
        out.append([c.strip() for c in row])
    return out


def _price_item_row(row: list[str], current_group: str) -> tuple[dict, str]:
    wg = row[0] if row and row[0] else current_group
    next_group = wg if row and row[0] else current_group
    item = {
        "work_group": wg,
        "no": row[1] if len(row) > 1 else "",
        "name": row[2] if len(row) > 2 else "",
        "manufacturer": row[3] if len(row) > 3 else "",
        "model": row[4] if len(row) > 4 else "",
        "spec": row[5] if len(row) > 5 else "",
        "unit": row[6] if len(row) > 6 else "",
        "qty": _coerce_number(row[7]) if len(row) > 7 else None,
        "repeat": _coerce_number(row[8]) if len(row) > 8 else None,
    }
    if is_total_row(item):
        item["unit_price"] = 0
        item["amount"] = 0
        item["price_source"] = "total_row_skipped"
        return item, next_group
    qty = item["qty"] if item["qty"] is not None else 1.0
    unit_price, source, conf = lookup_unit_price(item)
    item["unit_price"] = unit_price
    item["amount"] = round(float(qty) * float(unit_price), 0)
    item["price_source"] = source
    item["price_confidence"] = conf
    return item, next_group


def _write_header_row(ws, r: int) -> None:
    for c, (name, width) in enumerate(OUT_COLS, start=1):
        cell = ws.cell(row=r, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, size=10)
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[r].height = 22


def _write_item_row(ws, r: int, item: dict) -> None:
    values = [
        item.get("work_group", ""),
        item.get("no", ""),
        item.get("name", ""),
        item.get("manufacturer", ""),
        item.get("model", ""),
        item.get("spec", ""),
        item.get("unit", ""),
        item.get("qty"),
        item.get("repeat"),
        item.get("unit_price") or 0,
        item.get("amount") or 0,
        item.get("price_source", ""),
    ]
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        if c in (2, 7, 9):
            cell.alignment = CENTER
        elif c in (8, 10, 11):
            cell.alignment = RIGHT
            if isinstance(v, (int, float)):
                cell.number_format = NUM_FMT
        else:
            cell.alignment = LEFT
        cell.font = Font(size=10)


def _write_subtotal_row(ws, r: int, work_group: str, subtotal: float) -> None:
    ws.cell(row=r, column=1, value=work_group).font = Font(bold=True)
    cell_label = ws.cell(row=r, column=2, value="소계")
    cell_label.font = Font(bold=True)
    cell_label.alignment = CENTER
    cell_amt = ws.cell(row=r, column=11, value=int(subtotal))
    cell_amt.font = Font(bold=True)
    cell_amt.number_format = NUM_FMT
    cell_amt.alignment = RIGHT
    for c in range(1, len(OUT_COLS) + 1):
        ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
        ws.cell(row=r, column=c).border = BORDER


def _write_total_row(ws, r: int, total: float, label: str = "합 계") -> None:
    cell_label = ws.cell(row=r, column=3, value=label)
    cell_label.font = Font(bold=True, size=12)
    cell_label.alignment = CENTER
    cell_amt = ws.cell(row=r, column=11, value=int(total))
    cell_amt.font = Font(bold=True, size=12)
    cell_amt.number_format = NUM_FMT
    cell_amt.alignment = RIGHT
    for c in range(1, len(OUT_COLS) + 1):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 28


def _write_summary_sheet(wb: Workbook, sheets_summary: list[dict]) -> None:
    ws = wb.create_sheet("요약", 0)
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "공내역서 단가 대입 요약"
    cell.font = Font(bold=True, size=14)
    cell.alignment = CENTER
    cell.fill = HEADER_FILL

    headers = ["시트명", "항목 수", "단가 매핑", "스킵(합계 등)", "직접공사비 (₩)"]
    for c, h in enumerate(headers, start=1):
        x = ws.cell(row=3, column=c, value=h)
        x.fill = HEADER_FILL
        x.font = Font(bold=True)
        x.alignment = CENTER
        x.border = BORDER

    r = 4
    grand_total = 0
    for s in sheets_summary:
        for c, v in enumerate([
            s["sheet"], s["item_count"], s["priced_count"], s["skipped_count"], int(s["subtotal"]),
        ], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = CENTER if c <= 4 else RIGHT
            if c == 5:
                cell.number_format = NUM_FMT
        grand_total += s["subtotal"]
        r += 1

    ws.cell(row=r, column=1, value="전체 직접공사비").font = Font(bold=True)
    cell_g = ws.cell(row=r, column=5, value=int(grand_total))
    cell_g.font = Font(bold=True, size=12)
    cell_g.number_format = NUM_FMT
    cell_g.alignment = RIGHT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).border = BORDER

    # 요율 적용 박스
    r += 3
    ws.cell(row=r, column=1, value="요율 적용 (참고)").font = Font(bold=True, size=12)
    r += 1
    rates = [
        ("직접공사비", grand_total, 1.0),
        ("간접비 (15%)", grand_total * 0.15, 0.15),
        ("안전관리비 (1.86%)", (grand_total * 1.15) * 0.0186, 0.0186),
        ("산재보험 (3.8%)", (grand_total * 1.15) * 0.038, 0.038),
    ]
    subtotal = sum(v for _, v, _ in rates)
    rates.append(("소계", subtotal, None))
    rates.append(("VAT (10%)", subtotal * 0.10, 0.10))
    rates.append(("합계 (VAT 포함)", subtotal * 1.10, None))
    for label, val, _rate in rates:
        ws.cell(row=r, column=1, value=label).alignment = LEFT
        c = ws.cell(row=r, column=5, value=int(val))
        c.number_format = NUM_FMT
        c.alignment = RIGHT
        is_emphasis = label in ("합계 (VAT 포함)",)
        if is_emphasis:
            ws.cell(row=r, column=1).font = Font(bold=True, size=12)
            ws.cell(row=r, column=5).font = Font(bold=True, size=12)
            for cc in range(1, 6):
                ws.cell(row=r, column=cc).fill = TOTAL_FILL
        r += 1

    widths = [22, 12, 12, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28


def write_priced_xlsx(xls_in: Path, xlsx_out: Path) -> dict:
    src = _open(xls_in)
    wb = Workbook()
    # 기본으로 생기는 'Sheet' 제거
    default_sheet = wb.active
    wb.remove(default_sheet)

    sheets_summary: list[dict] = []
    seen_sigs: set[str] = set()

    for sheet_name in src.sheet_names:
        rows = _read_raw_rows(xls_in, sheet_name, src.engine)

        # 시트 시그니처로 중복 시트 식별 (chipotle BOQ의 경우 sheet 2/3 동일)
        items_preview = []
        for r in rows[:30]:
            kind = _classify_row(r)
            if kind == "item":
                items_preview.append(f"{r[2] if len(r) > 2 else ''}:{r[7] if len(r) > 7 else ''}")
            if len(items_preview) >= 5:
                break
        sig = "|".join(items_preview) + f"|n={sum(1 for r in rows if _classify_row(r) == 'item')}"

        is_duplicate = bool(items_preview) and sig in seen_sigs
        if items_preview:
            seen_sigs.add(sig)

        # 시트 작성
        ws_title = sheet_name if not is_duplicate else f"{sheet_name}(중복)"
        ws = wb.create_sheet(ws_title)

        out_r = 1
        current_group = ""
        wg_subtotal: dict[str, float] = defaultdict(float)
        wg_order: list[str] = []
        sheet_total = 0.0
        item_count = 0
        priced_count = 0
        skipped_count = 0
        header_drawn = False
        last_wg_in_sheet: str | None = None

        for raw in rows:
            kind = _classify_row(raw)

            if kind == "header":
                # 본격 BOQ 헤더 — 우리 형식으로 작성
                _write_header_row(ws, out_r)
                header_drawn = True
                out_r += 1
                continue

            if kind == "blank" or kind == "noise":
                continue

            if kind == "summary":
                # 합계/공과잡비 등 — 메모만 남기고 스킵
                name = raw[2] if len(raw) > 2 else ""
                cell = ws.cell(row=out_r, column=3, value=f"[{name}]")
                cell.fill = SUMMARY_FILL
                cell.font = Font(italic=True, color="888888")
                out_r += 1
                continue

            if kind == "label":
                # 공종 라벨 — 스킵 (헤더는 위에서 그렸음, 직전 행에 메타 정보면 표시만)
                if not header_drawn:
                    name = raw[2] if len(raw) > 2 else ""
                    if name:
                        ws.cell(row=out_r, column=1, value=name).font = Font(italic=True, color="666666")
                        out_r += 1
                continue

            if kind == "item":
                if not header_drawn:
                    _write_header_row(ws, out_r)
                    header_drawn = True
                    out_r += 1

                if is_duplicate:
                    # 중복 시트는 단가/금액 빈 채로 항목만 표기
                    item, current_group = _price_item_row(raw, current_group)
                    item["unit_price"] = 0
                    item["amount"] = 0
                    item["price_source"] = "duplicate_sheet_skipped"
                    _write_item_row(ws, out_r, item)
                    out_r += 1
                    item_count += 1
                    skipped_count += 1
                    continue

                item, current_group = _price_item_row(raw, current_group)

                if item["price_source"] == "total_row_skipped":
                    _write_item_row(ws, out_r, item)
                    out_r += 1
                    skipped_count += 1
                    item_count += 1
                    continue

                # work_group 변경 감지 → 직전 group 소계 추가
                wg = item.get("work_group") or "(미분류)"
                if last_wg_in_sheet is not None and wg != last_wg_in_sheet:
                    _write_subtotal_row(ws, out_r, last_wg_in_sheet, wg_subtotal[last_wg_in_sheet])
                    out_r += 1

                _write_item_row(ws, out_r, item)
                out_r += 1
                item_count += 1
                priced_count += 1

                amount = item.get("amount") or 0
                wg_subtotal[wg] += amount
                if wg not in wg_order:
                    wg_order.append(wg)
                sheet_total += amount
                last_wg_in_sheet = wg

        # 마지막 group 소계 + 시트 전체 합계
        if last_wg_in_sheet is not None and not is_duplicate:
            _write_subtotal_row(ws, out_r, last_wg_in_sheet, wg_subtotal[last_wg_in_sheet])
            out_r += 1
            _write_total_row(ws, out_r, sheet_total, label=f"{sheet_name} 직접공사비 합계")
            out_r += 1

        # 첫 행 고정
        ws.freeze_panes = "A2"

        sheets_summary.append({
            "sheet": ws_title,
            "item_count": item_count,
            "priced_count": priced_count,
            "skipped_count": skipped_count,
            "subtotal": 0 if is_duplicate else sheet_total,
            "duplicate": is_duplicate,
        })

    _write_summary_sheet(wb, sheets_summary)

    xlsx_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_out)

    return {
        "xls_in": str(xls_in),
        "xlsx_out": str(xlsx_out),
        "sheets": sheets_summary,
        "grand_total_direct": sum(s["subtotal"] for s in sheets_summary),
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xls_in")
    ap.add_argument("xlsx_out")
    args = ap.parse_args()

    result = write_priced_xlsx(Path(args.xls_in), Path(args.xlsx_out))

    print(f"Saved: {result['xlsx_out']}")
    print(f"Grand total (직접공사비): ₩{result['grand_total_direct']:,}")
    for s in result["sheets"]:
        flag = " (DUP)" if s["duplicate"] else ""
        print(f"  - {s['sheet']}{flag}: items={s['item_count']} priced={s['priced_count']} skipped={s['skipped_count']} subtotal=₩{int(s['subtotal']):,}")


if __name__ == "__main__":
    main()
