"""
공내역서 단가대입 Excel 출력기.

boq_workbook.json + boq_pricing.py 매핑 결과를 받아
실제 공내역서 Excel 파일을 생성한다.

사용법:
    python -X utf8 scripts/estimation/boq_excel_export.py chipotle-shinsegae
"""
from __future__ import annotations
import sys, json, copy, os
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "scripts" / "estimation"))

from boq_pricing import price_boq_workbook, build_cost_estimate_from_boq, extract_actual_area


def export_boq_excel(slug: str) -> Path:
    sample_dir = ROOT / "data" / "estimation_samples" / slug
    out_dir = sample_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    wb_path = sample_dir / "boq_workbook.json"
    if not wb_path.exists():
        raise FileNotFoundError(f"boq_workbook.json not found: {wb_path}")

    with open(wb_path, encoding="utf-8") as f:
        wb = json.load(f)

    wb2 = copy.deepcopy(wb)
    priced = price_boq_workbook(wb2)
    area_result = extract_actual_area(wb2)
    actual_area = area_result.get("actual_area_m2")
    estimate = build_cost_estimate_from_boq(priced, actual_area)

    # 갱신된 cost_estimate 저장
    est_path = sample_dir / "cost_estimate.json"
    with open(est_path, "w", encoding="utf-8") as f:
        json.dump(estimate, f, ensure_ascii=False, indent=2)

    # docs/data 경량 번들 갱신
    docs_dir = ROOT / "docs" / "data" / "estimation"
    docs_dir.mkdir(parents=True, exist_ok=True)
    bundle_path = docs_dir / f"{slug}.json"
    bundle = {
        "slug": slug,
        "generated_at": estimate.get("method", ""),
        "actual_area_m2": actual_area,
        "actual_area_pyeong": estimate.get("actual_area_pyeong"),
        "raw_total_krw": estimate.get("raw_total_krw"),
        "per_pyeong_avg": estimate.get("per_pyeong_avg"),
        "confidence": estimate.get("confidence"),
        "note": estimate.get("note"),
        "total_min": estimate.get("total_min"),
        "total_avg": estimate.get("total_avg"),
        "total_max": estimate.get("total_max"),
        "gongje_breakdown": estimate.get("gongje_breakdown"),
        "missing_from_briefing": estimate.get("missing_from_briefing"),
        "confidence_counter": priced.get("price_meta", {}).get("confidence_counter"),
    }
    with open(bundle_path, "w", encoding="utf-8") as f:
        json.dump(bundle, f, ensure_ascii=False, indent=2)

    # Excel 출력
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    xwb = openpyxl.Workbook()
    xwb.remove(xwb.active)

    # 스타일 정의
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    total_fill = PatternFill("solid", fgColor="D6E4F0")
    total_font = Font(bold=True, size=10)
    normal_font = Font(size=9)
    wg_fill = PatternFill("solid", fgColor="BDD7EE")
    wg_font = Font(bold=True, size=9)
    low_fill = PatternFill("solid", fgColor="FFE0B2")
    center = Alignment(horizontal="center", vertical="center")
    right_a = Alignment(horizontal="right", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _cell(ws, row, col, value, font=None, fill=None, align=None, number_format=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        if number_format:
            c.number_format = number_format
        c.border = thin
        return c

    # ── 시트 1: 전체 내역서 ──────────────────────────────────────────────────
    ws1 = xwb.create_sheet("전체내역서_단가대입")
    headers = ["공종", "No", "품명", "규격", "단위", "수량", "단가", "금액", "신뢰도", "단가근거"]
    col_widths = [12, 5, 35, 20, 6, 8, 14, 16, 8, 25]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        _cell(ws1, 1, i, h, font=header_font, fill=header_fill, align=center)
        ws1.column_dimensions[get_column_letter(i)].width = w

    row = 2
    items_all = estimate.get("items", [])
    prev_wg = None
    for it in items_all:
        wg = it.get("공종", "")
        if wg != prev_wg:
            _cell(ws1, row, 1, f"▶ {wg}", font=wg_font, fill=wg_fill, align=center)
            for c in range(2, 11):
                ws1.cell(row=row, column=c).fill = wg_fill
                ws1.cell(row=row, column=c).border = thin
            ws1.row_dimensions[row].height = 16
            row += 1
            prev_wg = wg

        conf = it.get("신뢰도", "")
        row_fill = low_fill if "LOW" in str(conf) else None

        _cell(ws1, row, 1, wg, font=normal_font, fill=row_fill, align=center)
        _cell(ws1, row, 2, it.get("No") or it.get("no", ""), font=normal_font, fill=row_fill, align=center)
        _cell(ws1, row, 3, it.get("품명", ""), font=normal_font, fill=row_fill)
        _cell(ws1, row, 4, it.get("규격") or "", font=normal_font, fill=row_fill)
        _cell(ws1, row, 5, it.get("단위", ""), font=normal_font, fill=row_fill, align=center)
        qty = it.get("수량")
        _cell(ws1, row, 6, qty, font=normal_font, fill=row_fill, align=right_a,
              number_format="#,##0.00")
        unit_p = it.get("단가", 0) or 0
        amount = it.get("금액", 0) or 0
        _cell(ws1, row, 7, unit_p, font=normal_font, fill=row_fill, align=right_a,
              number_format="#,##0")
        _cell(ws1, row, 8, amount, font=normal_font, fill=row_fill, align=right_a,
              number_format="#,##0")
        _cell(ws1, row, 9, conf, font=normal_font, fill=row_fill, align=center)
        _cell(ws1, row, 10, it.get("단가출처", ""), font=normal_font, fill=row_fill)
        row += 1

    # 합계 행
    raw = estimate.get("raw_total_krw", 0)
    _cell(ws1, row, 1, "직접공사비 합계", font=total_font, fill=total_fill)
    for c in range(2, 8):
        ws1.cell(row=row, column=c).fill = total_fill
        ws1.cell(row=row, column=c).border = thin
    _cell(ws1, row, 8, raw, font=total_font, fill=total_fill, align=right_a,
          number_format="#,##0")
    for c in range(9, 11):
        ws1.cell(row=row, column=c).fill = total_fill
        ws1.cell(row=row, column=c).border = thin

    # ── 시트 2: 공사비 요약 ──────────────────────────────────────────────────
    ws2 = xwb.create_sheet("공사비요약")
    summary_headers = ["구분", "최소(×0.85)", "평균", "최대(야간+15%)"]
    for i, h in enumerate(summary_headers, 1):
        _cell(ws2, 1, i, h, font=header_font, fill=header_fill, align=center)
        ws2.column_dimensions[get_column_letter(i)].width = 20

    rows_data = [
        ("직접공사비", "직접공사비"),
        ("간접비(15%)", "간접비"),
        ("안전관리비(1.86%)", "안전관리비"),
        ("산재보험(3.8%)", "산재보험"),
        ("소계", "소계"),
        ("VAT(10%)", "VAT"),
        ("합계(VAT포함)", "합계_VAT포함"),
    ]
    t_min = estimate.get("total_min", {})
    t_avg = estimate.get("total_avg", {})
    t_max = estimate.get("total_max", {})
    for i, (label, key) in enumerate(rows_data, 2):
        is_total = key in ("소계", "합계_VAT포함")
        f = total_font if is_total else normal_font
        fi = total_fill if is_total else None
        _cell(ws2, i, 1, label, font=f, fill=fi)
        _cell(ws2, i, 2, t_min.get(key, 0), font=f, fill=fi, align=right_a, number_format="#,##0")
        _cell(ws2, i, 3, t_avg.get(key, 0), font=f, fill=fi, align=right_a, number_format="#,##0")
        _cell(ws2, i, 4, t_max.get(key, 0), font=f, fill=fi, align=right_a, number_format="#,##0")

    # 면적 정보
    ws2.cell(row=10, column=1).value = "실측 면적(㎡)"
    ws2.cell(row=10, column=2).value = actual_area
    ws2.cell(row=11, column=1).value = "실측 면적(평)"
    ws2.cell(row=11, column=2).value = estimate.get("actual_area_pyeong")
    ws2.cell(row=12, column=1).value = "평당 단가(원)"
    ws2.cell(row=12, column=2).value = estimate.get("per_pyeong_avg")

    # 현설 누락 항목
    ws2.cell(row=14, column=1).value = "▶ 현장설명회 누락 항목 (미반영)"
    ws2.cell(row=14, column=1).font = Font(bold=True, color="C00000")
    for i, item in enumerate(estimate.get("missing_from_briefing", []), 15):
        ws2.cell(row=i, column=1).value = f"• {item}"

    # ── 시트 3: 신뢰도 안내 ─────────────────────────────────────────────────
    ws3 = xwb.create_sheet("신뢰도안내")
    notes = [
        ["신뢰도", "의미"],
        ["🟡 MED", "name 키워드 또는 work_group+unit 매핑 — 한국 인테리어 평균단가(2024-25) 사용"],
        ["🔴 LOW", "unit 단위만으로 fallback — 시장 평균 참고치"],
        ["", ""],
        ["주의사항", ""],
        ["", "협력사 확정 단가가 아닙니다. 발주 전 실측·견적 재산출 필요."],
        ["", "신세계 입점 2공구 야간공사 할증 15% 포함 (최대 시나리오)."],
        ["", "현설 누락 10개 항목은 이 내역서에 미반영."],
    ]
    for r, row_vals in enumerate(notes, 1):
        for c, val in enumerate(row_vals, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 60

    excel_path = out_dir / f"{slug}_공내역서_단가대입.xlsx"
    xwb.save(excel_path)
    return excel_path


if __name__ == "__main__":
    slug = sys.argv[1] if len(sys.argv) > 1 else "chipotle-shinsegae"
    out = export_boq_excel(slug)
    print(f"저장 완료: {out}")
    import json, copy
    from pathlib import Path
    ROOT2 = Path(__file__).resolve().parents[2]
    wb_path = ROOT2 / "data" / "estimation_samples" / slug / "boq_workbook.json"
    with open(wb_path, encoding="utf-8") as f:
        wb = json.load(f)
    wb2 = copy.deepcopy(wb)
    sys.path.insert(0, str(ROOT2 / "scripts" / "estimation"))
    from boq_pricing import price_boq_workbook
    priced = price_boq_workbook(wb2)
    meta = priced.get("price_meta", {})
    print(f"신뢰도: {meta.get('confidence_counter')}")
    print(f"출처: {meta.get('source_counter')}")
