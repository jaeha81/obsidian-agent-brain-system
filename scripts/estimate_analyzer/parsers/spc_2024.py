"""
SPC 표준 양식(2024) 파서.
"전체내역서" 시트와 "직사입자재" 시트를 정규화된 row 리스트로 반환한다.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional


# 전체내역서 컬럼 순서 (좌→우)
# 일부 시트는 헤더 행 위치가 불규칙하므로 키워드 탐지로 매핑한다.
_COLUMN_KEYWORDS: Dict[str, List[str]] = {
    "location":        ["위치", "구역"],
    "name":            ["명칭"],
    "manufacturer":    ["제조사"],
    "model_no":        ["모델", "model"],
    "spec":            ["규격"],
    "unit":            ["단위"],
    "quantity":        ["수량"],
    "loss_rate":       ["loss", "loss율"],
    "material_price":  ["자재비단가", "자재단가"],
    "material_amount": ["자재비금액", "자재금액"],
    "labor_price":     ["인건비단가"],
    "labor_amount":    ["인건비금액"],
    "total":           ["합계"],
    "remark":          ["비고"],
}

# 7대 공종 헤더 키워드
_TRADE_KEYWORDS = [
    "가설공사", "철거공사", "바닥공사", "벽체공사",
    "천정공사", "집기공사", "기타공사",
]

# 17종 작업조 키워드 (D6 analyzer 에서도 재사용)
CREW_LABELS = [
    "목공", "직영", "철거", "타일", "타일/부자재", "타일/메지",
    "습식", "습식-1", "경량", "도장", "소방", "구매", "청소",
    "폐기물", "금속", "간판", "전기",
]


def _safe_float(value: Any) -> float:
    """셀 값을 float으로 변환. 변환 불가 또는 빈 값은 0.0."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _clean_model_no(value: Any) -> Optional[str]:
    """모델No. 앞뒤 공백 제거, 빈 값은 None."""
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned if cleaned else None


def _detect_header_row(sheet) -> Optional[Dict[str, int]]:
    """
    시트를 스캔하여 컬럼 키워드가 가장 많이 매칭되는 행을 헤더 행으로 채택한다.
    반환: {필드명: 컬럼 인덱스(0-based)} 또는 None
    """
    best_row_idx = None
    best_mapping: Dict[str, int] = {}
    best_score = 0

    for row_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
        mapping: Dict[str, int] = {}
        for col_idx, cell_val in enumerate(row):
            if cell_val is None:
                continue
            cell_str = str(cell_val).strip().lower()
            for field, keywords in _COLUMN_KEYWORDS.items():
                if field in mapping:
                    continue
                for kw in keywords:
                    if kw.lower() in cell_str:
                        mapping[field] = col_idx
                        break
        if len(mapping) > best_score:
            best_score = len(mapping)
            best_row_idx = row_idx
            best_mapping = mapping

    return best_mapping if best_score >= 3 else None  # 최소 3개 컬럼 매칭


def _is_trade_header(row_values: tuple) -> Optional[str]:
    """행 값들 중 공종 키워드가 포함된 경우 공종 이름 반환."""
    for val in row_values:
        if val is None:
            continue
        for kw in _TRADE_KEYWORDS:
            if kw in str(val):
                return kw
    return None


def _parse_sheet(sheet, sheet_name: str) -> List[Dict]:
    """
    단일 시트를 정규화 row 리스트로 변환.
    헤더 탐지 실패 시 빈 리스트 반환(graceful).
    """
    col_map = _detect_header_row(sheet)
    if not col_map:
        return []

    # 헤더 행 다음 행부터 데이터
    header_row_idx = None
    for row_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
        cell_str_list = [str(v).strip().lower() for v in row if v is not None]
        match_count = sum(
            any(kw.lower() in cs for kw in kws for cs in cell_str_list)
            for kws in _COLUMN_KEYWORDS.values()
        )
        if match_count >= 3:
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        return []

    rows: List[Dict] = []
    current_trade = None

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_idx <= header_row_idx:
            continue

        row_vals = tuple(row)

        # 공종 그룹 헤더 행 감지
        trade = _is_trade_header(row_vals)
        if trade:
            current_trade = trade
            continue

        # 모든 셀이 비어있으면 스킵
        non_empty = [v for v in row_vals if v is not None and str(v).strip()]
        if not non_empty:
            continue

        def get(field: str) -> Any:
            idx = col_map.get(field)
            if idx is None or idx >= len(row_vals):
                return None
            return row_vals[idx]

        record: Dict[str, Any] = {
            "sheet":            sheet_name,
            "row_index":        row_idx,
            "trade_category":   current_trade,
            "location":         str(get("location") or "").strip() or None,
            "name":             str(get("name") or "").strip() or None,
            "manufacturer":     str(get("manufacturer") or "").strip() or None,
            "model_no":         _clean_model_no(get("model_no")),
            "spec":             str(get("spec") or "").strip() or None,
            "unit":             str(get("unit") or "").strip() or None,
            "quantity":         _safe_float(get("quantity")),
            "loss_rate":        _safe_float(get("loss_rate")),
            "material_price":   _safe_float(get("material_price")),
            "material_amount":  _safe_float(get("material_amount")),
            "labor_price":      _safe_float(get("labor_price")),
            "labor_amount":     _safe_float(get("labor_amount")),
            "total":            _safe_float(get("total")),
            "remark":           str(get("remark") or "").strip() or None,
        }
        rows.append(record)

    return rows


class SPCEstimateParser:
    """SPC 표준 견적 양식(2024) .xlsx / .xls 파서."""

    def parse(self, filepath: str) -> Dict:
        """
        반환:
          {
            "project_name": str,
            "rows": List[dict],       # 전체내역서 정규화 행
            "rows_direct": List[dict],  # 직사입자재 행
            "summary": {
              "total_material": float,
              "total_labor": float,
              "total_sum": float
            }
          }
        """
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            wb = self._open_xlsx(path)
        elif suffix in (".xls",):
            wb = self._open_xls_as_xlsx(path)
        else:
            raise ValueError(f"지원하지 않는 파일 형식: {suffix} (xlsx 또는 xls만 허용)")

        sheet_names = wb.sheetnames

        # 전체내역서 시트 탐색 (이름에 "전체내역서" 포함)
        main_sheet = self._find_sheet(wb, sheet_names, ["전체내역서", "전체 내역서", "내역서"])
        direct_sheet = self._find_sheet(wb, sheet_names, ["직사입자재", "직사입", "직사입 자재"])

        rows = _parse_sheet(main_sheet, "전체내역서") if main_sheet else []
        rows_direct = _parse_sheet(direct_sheet, "직사입자재") if direct_sheet else []

        # 프로젝트명: 표지 시트 또는 파일명에서
        project_name = self._extract_project_name(wb, sheet_names, path)

        total_material = sum(r["material_amount"] for r in rows)
        total_labor = sum(r["labor_amount"] for r in rows)
        total_sum = sum(r["total"] for r in rows)

        return {
            "project_name": project_name,
            "rows": rows,
            "rows_direct": rows_direct,
            "summary": {
                "total_material": total_material,
                "total_labor": total_labor,
                "total_sum": total_sum,
            },
        }

    # ── 내부 헬퍼 ──────────────────────────────────────────

    @staticmethod
    def _open_xlsx(path: Path):
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl 패키지가 필요합니다: pip install openpyxl") from exc
        return openpyxl.load_workbook(path, data_only=True)

    @staticmethod
    def _open_xls_as_xlsx(path: Path):
        """xlrd로 .xls 읽기. openpyxl Workbook 호환 래퍼 반환."""
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError("xlrd 패키지가 필요합니다: pip install xlrd") from exc

        # xlrd -> openpyxl 래핑 (최소 인터페이스만 제공)
        xls_wb = xlrd.open_workbook(str(path))

        class _XlrdSheet:
            def __init__(self, xls_sheet):
                self._s = xls_sheet
                self.title = xls_sheet.name

            def iter_rows(self, max_row=None, values_only=True):
                nrows = self._s.nrows
                if max_row is not None:
                    nrows = min(nrows, max_row)
                for r in range(nrows):
                    yield tuple(
                        self._s.cell_value(r, c) or None
                        for c in range(self._s.ncols)
                    )

        class _XlrdWorkbook:
            def __init__(self, xls_wb):
                self._wb = xls_wb
                self.sheetnames = [s.name for s in xls_wb.sheets()]
                self._sheets = {s.name: _XlrdSheet(s) for s in xls_wb.sheets()}

            def __getitem__(self, name):
                return self._sheets[name]

        return _XlrdWorkbook(xls_wb)

    @staticmethod
    def _find_sheet(wb, sheet_names: List[str], candidates: List[str]):
        """후보 이름 목록과 가장 잘 매칭되는 시트 반환. 없으면 None."""
        for candidate in candidates:
            for name in sheet_names:
                if candidate in name:
                    return wb[name]
        return None

    @staticmethod
    def _extract_project_name(wb, sheet_names: List[str], path: Path) -> str:
        # 표지 시트에서 첫 번째 비어있지 않은 셀 값 사용
        cover_keywords = ["표지", "전체표지", "cover"]
        for kw in cover_keywords:
            for name in sheet_names:
                if kw in name.lower():
                    try:
                        sheet = wb[name]
                        for row in sheet.iter_rows(max_row=10, values_only=True):
                            for val in row:
                                if val and str(val).strip():
                                    text = str(val).strip()
                                    # 너무 짧거나 숫자만인 값은 스킵
                                    if len(text) >= 3 and not re.match(r"^\d+$", text):
                                        return text
                    except Exception:
                        pass
        # 표지 탐색 실패 시 파일명 사용
        return path.stem
